# Импорты
import time
from tkinter import Tk
from tkinter.constants import X
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl.styles import PatternFill
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#Указание путя до списка моделей
Tk().withdraw()
filename = askopenfilename()
print("Файл найден")
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#Запись моделей в массив model_list
model_list = []
with open(filename, "r") as f:
    while True:
        line = f.readline()
        if line:
            model_list.append(line.replace("\n", ""))
        if not line:
            break
print("Модели записанны")
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#Форматирование названий моделей
for item in range(len(model_list)):
    if " " in model_list[item]:
        model_list[item] = model_list[item].replace(" ", "+")
print("Модели отфарматированны")
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


# Создание цветов exel
Green = PatternFill(start_color="ADFF2F", end_color="ADFF2F", fill_type="solid") # Зеленый
Red = PatternFill(start_color="B22222", end_color="B22222", fill_type="solid") # Красный
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


# Создание excel файла
book = openpyxl.Workbook()
sheet = book.active
sheet["A1"] = "Название"
sheet["C1"] = "Описание"
sheet["E1"] = "Характеристики"
sheet["G1"] = "Изображения"
xl_counter = 2
model_not_find = []
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#Указание настроек драйвера
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
options.headless = True
#options.add_argument('--disable-gpu')
#options.add_argument('--remote-debugging-port=9222')
options.add_argument('--enable-javascript')
options.add_argument('--user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36"')
#options.add_argument('--no-sandbox')
#options.add_argument('--ignore-certificate-errors')
#options.add_argument('--allow-insecure-localhost')
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


# Переход по ссылкам
for model_index in range(len(model_list)):
    print(f"{len(model_list)} : {model_index + 1}")
    print(model_list[model_index])
    print()
    start_page = f"https://www.dns-shop.ru/search/?q={model_list[model_index]}"
    browser = webdriver.Chrome("chromedriver.exe", options=options)
    browser.set_window_size(1920, 1080)
    browser.get(start_page)
    time.sleep(2)
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


    

    # Проверка на существование
    if browser.current_url == start_page:
        print(f"Модель {model_list[model_index]} не найдена!")
        browser.close()
        print()
        model_not_find.append(model_list[model_index])
        continue
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    

    # Поиск названия товара
    while True:
        try:
            browser.find_element_by_class_name('product-card-top__title')
        except Exception as e:
            time.sleep(1)
            continue
        title = browser.find_element_by_class_name('product-card-top__title').text
        break

    sheet[f"A{xl_counter}"] = title
    sheet[f"A{xl_counter}"].fill = Green
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


    # Перемещение курсора
    action = ActionChains(browser)
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    
    
    # Поиск и переход по изображеням
    images_list = []
    image_slider_items = browser.find_elements_by_class_name('product-images-slider__img')
    for image in range(len(image_slider_items)):
        action.move_to_element(browser.find_elements_by_class_name('product-images-slider__img')[image]).perform()
        if len(images_list) > 3:
            break
        else:
            if browser.find_element_by_class_name('product-images-slider__main-img').get_attribute('src') not in images_list:
                images_list.append(browser.find_element_by_class_name('product-images-slider__main-img').get_attribute('src'))
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


    # Запись ссылок на изображения
    images = ""
    for item in images_list:
        images += f"{item}; "

    sheet[f"G{xl_counter}"] = images
    sheet[f"G{xl_counter}"].fill = Green

    images = ""
    images_list.clear()
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


    # Поиск описания
    while True:
        try:
            browser.find_element_by_class_name('product-card-tabs__title')
        except Exception as e:
            time.sleep(1)
            continue
        
        browser.find_element_by_class_name('product-card-tabs__title-icon_description').click()
        break

    time.sleep(2)

    description = browser.find_element_by_xpath("/html/body/div[1]/div[4]/div[2]/div[1]/div[1]/p").text

    sheet[f"C{xl_counter}"] = description
    sheet[f"C{xl_counter}"].fill = Green
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------\
    

    # Поиск Тех. характеристик
    last_url = browser.current_url
    while True:
        try:
            browser.find_element_by_class_name('product-card-tabs__title-icon_characteristics')
        except Exception as e:
            time.sleep(1)
            continue

        browser.find_element_by_class_name('product-card-tabs__title-icon_characteristics').click()
        if browser.current_url.split('/')[-2] != 'characteristics':
            continue
        else:
            break

        
    time.sleep(2)
    rows = [item.text for item in browser.find_elements_by_tag_name('td')]
    
    characters = ""
    for row in rows:
        characters += " " + str(row)


    sheet[f"E{xl_counter}"] = characters
    sheet[f"E{xl_counter}"].fill = Green
    characters = ""
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------


    # Закрытие окна, инкреметнация ячейки
    xl_counter += 1
    browser.close()
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Внесение ненайденных моделей
if len(model_not_find) > 0:
    for item in model_not_find:
        sheet[f"A{xl_counter + 5}"] = f"{item} Не найдено"
        sheet[f"A{xl_counter + 5}"].fill = Red
        xl_counter += 1
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------\


# Cохранение файла по адресу
Tk().withdraw()
filepath = askdirectory()
book.save(f"{filepath}\pars_data.xlsx")
book.close()
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------\


browser.quit()